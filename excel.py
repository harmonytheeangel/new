import openpyxl
from openpyxl.styles import PatternFill
sheet = openpyxl.load_workbook("core_n.xlsx")
inv = sheet["Sheet1"]

maxrow = inv.max_row

#print(sheet.sheetnames)

value1 = inv["E3" ].value
value2 = inv["C6"].value
#print(value1 , value2)

#second option
# for i in range(1 ,maxrow  + 1):
#      value = inv.cell(i, maxrow).value
#      print(value)

#print(sheet.active.title) --to get active worksheet

#print(inv["A1"].value , )
#print(inv.cell(1 ,2 ).value)
maxcolumn = inv.max_column
for a in range(8 , maxrow + 1):
    print("\n "
    )
    for b in range(1 , maxcolumn + 1):
            print(f"{inv.cell(a, b).value}")#to print all the values in the file
        #print(inv.cell(a , b).value)
#print(inv.cell(1 , 5).value)
inv["C2"].fill = PatternFill("solid", fgColor= "33FF42")
inv["C3"].fill = PatternFill("solid", fgColor= "33FF42")
inv["C4"].fill = PatternFill("solid", fgColor= "33FF42")
inv["C5"].fill = PatternFill("solid", fgColor= "33FF42")
inv["C6"].fill = PatternFill("solid", fgColor= "33FF42")
inv["C7"].fill = PatternFill("solid", fgColor= "FF5533")
inv.cell(row= 7 , column= 1 , value= "Samuel")
inv.cell(row= 7 , column= 2 , value= "David")
inv.cell(row= 7 , column= 3 , value= 19)
inv.cell(row= 7 , column= 4 , value= "19/5/2005")
inv.cell(row= 7 , column= 5 , value= "Abuja")
inv.cell(row= 7 , column= 6 , value= 32)

sheet.save("core_n.xlsx")